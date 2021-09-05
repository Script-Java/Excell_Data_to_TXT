import tkinter.messagebox
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import workbook

root = Tk()
root.title("Excell Data Converter")
blue = "#8B98F7"
root.config(bg=blue)



apps = []


def upload_file():
    global filename
    filename = filedialog.askopenfilename(initialdir="C:/", title="select file",
                                          filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
    apps.append(filename)
    for app in apps:
        label = Label(fileFrame, text=app, font=('Arial',16))
        label.pack()


def node_tab():
    try:
        wb = load_workbook(filename)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        columnC = ws["C"]
        columnD = ws["D"]
    except InvalidFileException:
        tkinter.messagebox.showerror(title="Invalid File Type",message="Please make sure to use Excell File or (.xlxs)")
    for cell in columnC:
        if cell.value != 0 and isinstance(cell.value, int):
            outputText.insert(END, cell.value)
            outputText.insert(END, "\n")

    for cell in columnD:
        if cell.value != 0 and isinstance(cell.value, int):
            outputText.insert(END, "-" + str(cell.value))
            outputText.insert(END, "\n")
    global outputValue
    outputValue = outputText.get("1.0", END)
    print(outputValue)


def save_file():
    try:
        file = open("output.txt", "w")
        file.write(outputValue)
        tkinter.messagebox.showinfo(title="File Saved", message="You're File has been saved")
        root.destroy()
    except NameError:
        tkinter.messagebox.showerror(title="Unknown Output",message="Please Run The File")


def run_file():
    node_tab()


# Tkinter structure
heading = Label(root,
                text="Excell Data Converter",
                font=('Arial', 26),
                bg=blue
                )

dataLabelFrame = LabelFrame(root,
                            text="Upload Data",
                            font=('Arial', 18),
                            bg=blue
                            )

fileFrame = Frame(dataLabelFrame,
                  width=600,
                  height=300,
                  )

outputLabelFrame = LabelFrame(root,
                              text="Output",
                              font=('Arial', 18),
                              bg=blue
                              )

outputText = Text(outputLabelFrame,
                  bg=blue,
                  font=('Arial',14)
                  )

uploadButton = Button(root,
                      text="Upload File",
                      width=16,
                      height=2,
                      font=('Arial', 12),
                      command=upload_file
                      )

runButton = Button(root,
                   text="Run File",
                   width=16,
                   height=2,
                   font=('Arial', 12),
                   command=run_file
                   )
saveButton = Button(root,
                    text="Save File",
                    width=16,
                    height=2,
                    font=('Arial', 12),
                    command=save_file
                    )
heading.grid(row=0, column=0)
dataLabelFrame.grid(row=1, column=0,pady=15)
fileFrame.grid(row=0, column=0)
outputLabelFrame.grid(row=2, column=0)
outputText.grid(row=0, column=0)
uploadButton.grid(row=3, column=0, pady=5)
runButton.grid(row=4, column=0, pady=5)
saveButton.grid(row=5, column=0, pady=5)
fileFrame.pack_propagate(False)

root.mainloop()
